import os
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from server import db
from server import reports


class TestReportsRegistry(unittest.TestCase):
    def setUp(self) -> None:
        self.conn = db.get_connection(":memory:")
        self.tmp = tempfile.TemporaryDirectory()

    def tearDown(self) -> None:
        self.conn.close()
        self.tmp.cleanup()

    def test_list_reports_defaults_to_last_five(self) -> None:
        for i in range(7):
            db.create_report(
                self.conn,
                report_id=f"rpt_{i}",
                job_id=None,
                pair_id=None,
                source_dataset=f"src_{i}",
                target_dataset=f"tgt_{i}",
                file_path=f"/tmp/r{i}.xlsx",
                file_name=f"r{i}.xlsx",
                summary={"n": i},
            )

        rows_default = db.list_reports(self.conn)
        rows_all = db.list_reports(self.conn, limit=50)
        self.assertEqual(len(rows_default), 5)
        self.assertEqual(len(rows_all), 7)

    def test_export_query_registers_report_and_returns_report_id(self) -> None:
        sql = "SELECT id, name FROM source_customers ORDER BY id"
        with patch("server.reports._reports_dir", return_value=self.tmp.name):
            result = reports.export_query_to_xlsx(
                headers=["id", "name"],
                rows=[["1", "Alice"], ["2", "Bob"]],
                filename="query_test.xlsx",
                sql_query=sql,
                conn=self.conn,
            )

        self.assertIn("report_id", result)
        self.assertTrue(result["report_id"])
        self.assertTrue(os.path.exists(result["file_path"]))
        self.assertEqual(result["row_count"], 2)

        stored = db.get_report(self.conn, result["report_id"])
        self.assertIsNotNone(stored)
        self.assertEqual(stored["source_dataset"], "query_export")
        self.assertEqual(stored["target_dataset"], "query_export")

        wb = load_workbook(result["file_path"])
        self.assertIn("SQL", wb.sheetnames)
        self.assertEqual(wb["SQL"]["A1"].value, "SQL Query")
        self.assertEqual(wb["SQL"]["A2"].value, sql)

    def test_export_column_summary_registers_report_and_returns_report_id(self) -> None:
        summary_result = {
            "dataset": "source_sample",
            "summaries": [
                {
                    "column": "name",
                    "top_values": [{"value": "Alice", "count": 1}],
                    "blank_or_null_count": 0,
                }
            ],
        }
        with patch("server.reports._reports_dir", return_value=self.tmp.name):
            result = reports.export_column_summary_to_xlsx(
                summary_result=summary_result,
                top_n=5,
                filename="colsum_test.xlsx",
                conn=self.conn,
            )

        self.assertIn("report_id", result)
        self.assertTrue(result["report_id"])
        self.assertTrue(os.path.exists(result["file_path"]))
        self.assertEqual(result["dataset"], "source_sample")

        stored = db.get_report(self.conn, result["report_id"])
        self.assertIsNotNone(stored)
        self.assertEqual(stored["source_dataset"], "source_sample")
        self.assertEqual(stored["target_dataset"], "source_sample")

    def test_reports_dir_uses_report_folder_meta_when_set(self) -> None:
        custom_dir = os.path.join(self.tmp.name, "custom_reports")
        db.set_meta(self.conn, "report_folder", custom_dir)

        resolved = reports._reports_dir(self.conn)

        self.assertEqual(resolved, custom_dir)
        self.assertTrue(os.path.isdir(custom_dir))


if __name__ == "__main__":
    unittest.main()
