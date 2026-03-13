import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from server import db
from server import jobs
from server import reports


class _FakeCursor:
    def __init__(self, rows):
        self._rows = list(rows)
        self._pos = 0

    def fetchmany(self, size=1):
        if self._pos >= len(self._rows):
            return []
        batch = self._rows[self._pos : self._pos + size]
        self._pos += len(batch)
        return batch


class TestExportQueryPerformance(unittest.TestCase):
    def setUp(self) -> None:
        self.conn = db.get_connection(":memory:")
        self.tmp = tempfile.TemporaryDirectory()

    def tearDown(self) -> None:
        self.conn.close()
        self.tmp.cleanup()

    def test_export_query_streams_from_cursor_like_rows(self) -> None:
        cursor = _FakeCursor([("1", "Alice"), ("2", "Bob"), ("3", "Carol")])
        with patch("server.reports._reports_dir", return_value=self.tmp.name):
            result = reports.export_query_to_xlsx(
                headers=["id", "name"],
                rows=cursor,
                filename="query_stream.xlsx",
                sql_query="SELECT id, name FROM source_customers",
                conn=self.conn,
            )

        self.assertEqual(result["row_count"], 3)
        wb = load_workbook(result["file_path"], read_only=True)
        rows = list(wb["Results"].iter_rows(values_only=True))
        wb.close()
        self.assertEqual(rows[0], ("id", "name"))
        self.assertEqual(rows[1], ("1", "Alice"))
        self.assertEqual(rows[3], ("3", "Carol"))

    def test_start_export_query_job_sync_updates_progress(self) -> None:
        with patch("server.jobs.db.list_datasets", return_value=[{"id": "source_sample"}]), patch(
            "server.jobs._execute_query_export", return_value={"report_id": "r1", "row_count": 7}
        ):
            result = jobs.start_export_query_job(
                sql="SELECT 1",
                filename="out.xlsx",
                conn=self.conn,
            )

        self.assertEqual(result["state"], "succeeded")
        self.assertEqual(result["progress"]["row_count"], 7)
        job = db.get_job(self.conn, result["job_id"])
        self.assertIsNotNone(job)
        self.assertEqual(job["state"], "succeeded")
        self.assertEqual(job["progress"]["row_count"], 7)


if __name__ == "__main__":
    unittest.main()
