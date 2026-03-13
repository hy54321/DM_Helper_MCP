"""
DuckDB Query Engine.

Manages ephemeral in-memory connections, view registration from catalog
datasets, and safe query execution.  Ported from
``workspace/ui/sql_query.py`` patterns.
"""

from __future__ import annotations

import csv
import os
import re
from contextlib import contextmanager
from typing import Any, Dict, List, Optional, Tuple


# ═══════════════════════════════════════════════════════════════
#  Name sanitisation
# ═══════════════════════════════════════════════════════════════

def sanitize_name(raw: str) -> str:
    """Turn a raw file/sheet/column name into a safe SQL identifier."""
    name = re.sub(r"[^\w]", "_", raw)
    name = re.sub(r"_+", "_", name).strip("_")
    if name and name[0].isdigit():
        name = f"t_{name}"
    return name or "unnamed"


def sanitize_column_names(
    columns: List[str],
) -> Tuple[List[str], Dict[str, str]]:
    """Sanitise a list of raw column headers.

    Returns ``(safe_names, column_map)`` where *column_map* maps each
    original name to its safe equivalent.
    """
    safe_names: List[str] = []
    column_map: Dict[str, str] = {}
    seen: Dict[str, int] = {}
    for raw in columns:
        safe = sanitize_name(str(raw).strip()) if raw else "unnamed"
        if safe in seen:
            seen[safe] += 1
            safe = f"{safe}_{seen[safe]}"
        else:
            seen[safe] = 1
        safe_names.append(safe)
        column_map[str(raw)] = safe
    return safe_names, column_map


# ═══════════════════════════════════════════════════════════════
#  SQL helpers
# ═══════════════════════════════════════════════════════════════

def quote(name: str) -> str:
    """Double-quote a SQL identifier."""
    return f'"{name}"'


def esc(value: str) -> str:
    """Escape a value for embedding inside a SQL string literal."""
    return value.replace("'", "''")


# ═══════════════════════════════════════════════════════════════
#  Header reading
# ═══════════════════════════════════════════════════════════════

def detect_text_encoding(path: str) -> Tuple[str, str]:
    """Detect a text file's encoding for Python and DuckDB CSV reads.

    Returns ``(python_encoding, duckdb_encoding)``.
    """
    try:
        with open(path, "rb") as fh:
            sample = fh.read(4096)
    except Exception:
        return ("utf-8-sig", "utf-8")

    if sample.startswith(b"\xef\xbb\xbf"):
        return ("utf-8-sig", "utf-8")
    if sample.startswith(b"\xff\xfe") or sample.startswith(b"\xfe\xff"):
        return ("utf-16", "utf-16")

    if sample:
        even_nuls = sample[::2].count(0)
        odd_nuls = sample[1::2].count(0)
        threshold = max(8, len(sample) // 16)
        if odd_nuls >= threshold and odd_nuls > even_nuls * 2:
            return ("utf-16le", "utf-16")
        if even_nuls >= threshold and even_nuls > odd_nuls * 2:
            return ("utf-16be", "utf-16")

    return ("utf-8-sig", "utf-8")


def read_csv_headers(path: str, encoding: Optional[str] = None) -> List[str]:
    """Read the first row of a CSV to extract column headers."""
    csv_py_encoding = encoding or detect_text_encoding(path)[0]
    try:
        with open(path, newline="", encoding=csv_py_encoding) as fh:
            sample = fh.read(8192)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ","
                if ";" in sample and "," not in sample:
                    delimiter = ";"
            reader = csv.reader(fh, delimiter=delimiter)
            return next(reader, [])
    except Exception:
        return []


def count_csv_rows(path: str, encoding: Optional[str] = None) -> Optional[int]:
    """Count data rows in a CSV-like file (excluding the header row)."""
    csv_py_encoding = encoding or detect_text_encoding(path)[0]
    try:
        with open(path, newline="", encoding=csv_py_encoding) as fh:
            sample = fh.read(8192)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ","
                if ";" in sample and "," not in sample:
                    delimiter = ";"
            reader = csv.reader(fh, delimiter=delimiter)
            total_rows = sum(1 for _ in reader)
            return max(total_rows - 1, 0)
    except Exception:
        return None


def read_excel_sheets(path: str) -> List[Tuple[str, List[str]]]:
    """Return ``[(sheet_name, [col_headers]), ...]`` for an Excel file."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        sheets: List[Tuple[str, List[str]]] = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            headers: List[str] = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(c) if c is not None else "" for c in row]
                break
            if headers:
                sheets.append((sname, headers))
        wb.close()
        return sheets
    except Exception:
        return []


def count_excel_sheet_rows(path: str, sheet_name: str) -> Optional[int]:
    """Count data rows in a specific Excel sheet (excluding the header row)."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell not in (None, "") for cell in row):
                count += 1
        wb.close()
        return count
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════
#  DuckDB connection factory
# ═══════════════════════════════════════════════════════════════

@contextmanager
def connect(datasets: List[Dict[str, Any]]):
    """Yield an in-memory DuckDB connection with views for *datasets*.

    Each dataset dict must contain at least::

        name, file_path, sheet_name, ext, columns, raw_columns, column_map
    """
    import duckdb

    conn = duckdb.connect()
    try:
        # Case-insensitive collation (best-effort, internal bypass of SqlGuard)
        try:
            conn.execute("SET default_collation='nocase'")
        except Exception:
            pass

        # Excel extension
        _load_excel_extension(conn)

        # Register views
        _register_views(conn, datasets)

        yield conn
    finally:
        conn.close()


def _load_excel_extension(conn) -> None:
    """Load the DuckDB excel extension (INSTALL if needed)."""
    try:
        conn.execute("LOAD 'excel'")
        return
    except Exception:
        pass
    try:
        conn.execute("INSTALL 'excel'")
        conn.execute("LOAD 'excel'")
    except Exception:
        pass  # CSV-only workflows still work


def _register_views(conn, datasets: List[Dict[str, Any]]) -> None:
    """Create one DuckDB view per dataset."""
    for ds in datasets:
        name = ds["name"] if "name" in ds else ds["id"]
        raw_cols: List[str] = ds.get("raw_columns") or []
        safe_cols: List[str] = ds.get("columns") or []
        raw_path = ds["file_path"].replace("\\", "/")
        path = esc(raw_path)
        sheet = esc(ds.get("sheet_name", ""))
        ext = ds.get("ext", "")

        # Pick the correct DuckDB table function
        if ext == ".csv":
            csv_duckdb_encoding = str(ds.get("csv_encoding", "") or "").strip().lower()
            if not csv_duckdb_encoding:
                _, csv_duckdb_encoding = detect_text_encoding(ds["file_path"])
            if csv_duckdb_encoding == "utf-16":
                src = (
                    "read_csv_auto("
                    f"'{path}', header=true, all_varchar=true, encoding='{csv_duckdb_encoding}'"
                    ")"
                )
            else:
                src = f"read_csv_auto('{path}', header=true, all_varchar=true)"
        elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            src = f"read_xlsx('{path}', sheet='{sheet}', header=true, all_varchar=true)"
        elif ext == ".xls":
            src = f"read_xls('{path}', sheet='{sheet}', header=true, all_varchar=true)"
        else:
            continue

        # Build aliased SELECT with COALESCE for NULLs
        sel = "*"
        if raw_cols and safe_cols and len(raw_cols) == len(safe_cols):
            parts = []
            for raw, safe in zip(raw_cols, safe_cols):
                parts.append(f"COALESCE({quote(str(raw))}, '') AS {safe}")
            sel = ", ".join(parts)

        view_name = quote(name)
        sql = f"CREATE OR REPLACE VIEW {view_name} AS SELECT {sel} FROM {src}"
        try:
            conn.execute(sql)
            continue
        except Exception:
            pass

        # Fallback: raw columns only
        try:
            conn.execute(
                f"CREATE OR REPLACE VIEW {view_name} AS SELECT * FROM {src}"
            )
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════════
#  Result formatting
# ═══════════════════════════════════════════════════════════════

def format_results(
    headers: List[str],
    rows: List[list],
    total: int,
    limit: int,
) -> str:
    """Format query results as a fixed-width text table."""
    if not rows:
        return "Query returned 0 rows."
    lines: List[str] = [f"Query returned {total} row(s)."]
    if total > limit:
        lines.append(f"Showing first {limit} rows.")
    lines.append("")

    # Column widths (capped at 50)
    widths = [len(str(h)) for h in headers]
    for row in rows[:limit]:
        for i, v in enumerate(row):
            if i < len(widths):
                widths[i] = max(widths[i], len(str(v) if v is not None else "NULL"))
    widths = [min(w, 50) for w in widths]

    hline = " | ".join(
        str(h).ljust(widths[i])[:widths[i]] for i, h in enumerate(headers)
    )
    lines.append(hline)
    lines.append("-" * len(hline))
    for row in rows[:limit]:
        rline = " | ".join(
            (str(v) if v is not None else "NULL").ljust(widths[i])[:widths[i]]
            for i, v in enumerate(row)
            if i < len(widths)
        )
        lines.append(rline)
    return "\n".join(lines)
