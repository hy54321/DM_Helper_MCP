"""
ReportService.

Writes XLSX comparison reports to disk and tracks report metadata
in SQLite.  Workbook tabs: Summary, Schema_Drift, Added, Removed, Changed.
"""

from __future__ import annotations

import os
import re
import sys
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, Iterator, List, Optional

from server import db


def _default_reports_dir() -> str:
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, "reports")


def _reports_dir(conn=None) -> str:
    custom_dir = ""
    own_conn = False
    if conn is None:
        try:
            conn = db.get_connection()
            own_conn = True
        except Exception:
            conn = None
    if conn is not None:
        try:
            custom_dir = (db.get_meta(conn, "report_folder", "") or "").strip()
        finally:
            if own_conn:
                conn.close()
    d = custom_dir or _default_reports_dir()
    os.makedirs(d, exist_ok=True)
    return d


# ═══════════════════════════════════════════════════════════════
#  Sheet naming helpers (Excel 31-char limit)
# ═══════════════════════════════════════════════════════════════

_INVALID_SHEET = re.compile(r"[\\/*?\[\]:]")
_INVALID_FILENAME = re.compile(r'[<>:"/\\|?*]')
_TIMESTAMP_SUFFIX = re.compile(r"_\d{8}_\d{6}$")
_DEFAULT_XLSX_STREAM_THRESHOLD = 20000


def _xlsx_stream_threshold() -> int:
    raw = (os.getenv("PROTOQUERY_XLSX_STREAM_THRESHOLD", "") or "").strip()
    if raw:
        try:
            value = int(raw)
            if value >= 1000:
                return value
        except ValueError:
            pass
    return _DEFAULT_XLSX_STREAM_THRESHOLD


def _iter_rows(rows: Any, batch_size: int = 2000) -> Iterator[Any]:
    """Iterate rows from a sequence or DB cursor-like object."""
    if hasattr(rows, "fetchmany"):
        while True:
            batch = rows.fetchmany(batch_size)
            if not batch:
                break
            for row in batch:
                yield row
        return

    for row in rows:
        yield row


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    name = _INVALID_SHEET.sub("_", name)
    return name[:max_len]


def _unique_sheet_names(names: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    result: List[str] = []
    for n in names:
        safe = _safe_sheet_name(n)
        if safe in seen:
            seen[safe] += 1
            suffix = f"_{seen[safe]}"
            safe = _safe_sheet_name(n, 31 - len(suffix)) + suffix
        else:
            seen[safe] = 1
        result.append(safe)
    return result


def _normalize_report_filename(
    filename: Optional[str],
    default_stem: str,
    default_ext: str = ".xlsx",
) -> str:
    """Build a safe report filename and ensure a timestamp suffix is present."""
    raw_name = (filename or "").strip() or f"{default_stem}{default_ext}"
    sanitized = _INVALID_FILENAME.sub("_", raw_name)
    stem, ext = os.path.splitext(sanitized)
    if not stem:
        stem = default_stem
    if not ext:
        ext = default_ext

    if not _TIMESTAMP_SUFFIX.search(stem):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = f"{stem}_{ts}"
    return f"{stem}{ext}"


# ═══════════════════════════════════════════════════════════════
#  XLSX writer
# ═══════════════════════════════════════════════════════════════

def _flatten_changed_rows(
    key_columns: List[str],
    compare_columns: List[str],
    changed_headers: List[str],
    changed_data: List[List[Any]],
    field_pairs: Optional[List[Dict[str, str]]] = None,
) -> tuple[List[str], List[List[Any]]]:
    """Convert changed rows from wide format to one row per changed field."""
    header_index = {h: i for i, h in enumerate(changed_headers)}
    key_indices = [header_index.get(k) for k in key_columns]
    value_pairs: List[tuple[str, int, int]] = []
    if field_pairs:
        for fp in field_pairs:
            label = fp.get("label") or f"{fp.get('source_field', '')}->{fp.get('target_field', '')}"
            source_idx = header_index.get(fp.get("source_header", ""))
            target_idx = header_index.get(fp.get("target_header", ""))
            if source_idx is not None and target_idx is not None:
                value_pairs.append((label, source_idx, target_idx))
    else:
        for col in compare_columns:
            source_idx = header_index.get(f"source_{col}")
            target_idx = header_index.get(f"target_{col}")
            if source_idx is not None and target_idx is not None:
                value_pairs.append((col, source_idx, target_idx))

    flat_headers = [*key_columns, "Field", "Target", "Source"]
    flat_data: List[List[Any]] = []
    for row in changed_data:
        key_values = [row[i] if i is not None else "" for i in key_indices]
        for field_name, source_idx, target_idx in value_pairs:
            source_val = row[source_idx]
            target_val = row[target_idx]
            if source_val != target_val:
                flat_data.append([*key_values, field_name, target_val, source_val])
    return flat_headers, flat_data


def _iter_flatten_changed_rows(
    key_columns: List[str],
    compare_columns: List[str],
    changed_headers: List[str],
    changed_data: List[List[Any]],
    field_pairs: Optional[List[Dict[str, str]]] = None,
):
    """Yield changed rows in one-row-per-field format."""
    header_index = {h: i for i, h in enumerate(changed_headers)}
    key_indices = [header_index.get(k) for k in key_columns]
    value_pairs: List[tuple[str, int, int]] = []
    if field_pairs:
        for fp in field_pairs:
            label = fp.get("label") or f"{fp.get('source_field', '')}->{fp.get('target_field', '')}"
            source_idx = header_index.get(fp.get("source_header", ""))
            target_idx = header_index.get(fp.get("target_header", ""))
            if source_idx is not None and target_idx is not None:
                value_pairs.append((label, source_idx, target_idx))
    else:
        for col in compare_columns:
            source_idx = header_index.get(f"source_{col}")
            target_idx = header_index.get(f"target_{col}")
            if source_idx is not None and target_idx is not None:
                value_pairs.append((col, source_idx, target_idx))

    for row in changed_data:
        key_values = [row[i] if i is not None else "" for i in key_indices]
        for field_name, source_idx, target_idx in value_pairs:
            source_val = row[source_idx]
            target_val = row[target_idx]
            if source_val != target_val:
                yield [*key_values, field_name, target_val, source_val]


def _write_comparison_report_streaming(
    comparison_result: Dict[str, Any],
    filename: Optional[str] = None,
    job_id: Optional[str] = None,
    pair_id: Optional[str] = None,
    conn=None,
) -> Dict[str, Any]:
    """Write comparison report in write-only mode for large datasets."""
    from openpyxl import Workbook

    source_id = comparison_result["source"]
    target_id = comparison_result["target"]
    key_columns = comparison_result.get("key_columns", [])
    compare_columns = comparison_result.get("compare_columns", [])
    schema_drift = comparison_result.get("schema_drift", {})
    added = comparison_result.get("added", {})
    removed = comparison_result.get("removed", {})
    changed = comparison_result.get("changed", {})

    added_headers = added.get("headers", [])
    added_data = added.get("data", [])
    removed_headers = removed.get("headers", [])
    removed_data = removed.get("data", [])
    changed_headers = changed.get("headers", [])
    changed_data = changed.get("data", [])
    changed_field_pairs = changed.get("field_pairs", [])

    added_count = len(added_data)
    removed_count = len(removed_data)
    changed_count = len(changed_data)

    filename = _normalize_report_filename(
        filename=filename,
        default_stem=f"comparison_{source_id}_vs_{target_id}",
    )
    file_path = os.path.join(_reports_dir(conn), filename)

    wb = Workbook(write_only=True)

    ws_sum = wb.create_sheet("Summary")
    summary_rows = [
        ("Comparison Report", ""),
        ("", ""),
        ("Source Dataset", source_id),
        ("Target Dataset", target_id),
        ("Key Columns", ", ".join(key_columns)),
        ("Compare Columns", ", ".join(compare_columns)),
        ("", ""),
        ("Timestamp", datetime.now(timezone.utc).isoformat()),
        ("", ""),
        ("Category", "Count"),
        ("Added (target-only)", added_count),
        ("Removed (source-only)", removed_count),
        ("Changed", changed_count),
        ("", ""),
        ("Schema Drift", ""),
        ("Source-only columns", ", ".join(schema_drift.get("source_only", []))),
        ("Target-only columns", ", ".join(schema_drift.get("target_only", []))),
    ]
    for row_data in summary_rows:
        ws_sum.append(list(row_data))

    ws_drift = wb.create_sheet("Schema_Drift")
    ws_drift.append(["Direction", "Column"])
    for c in schema_drift.get("source_only", []):
        ws_drift.append(["Source only", c])
    for c in schema_drift.get("target_only", []):
        ws_drift.append(["Target only", c])

    ws_added = wb.create_sheet("Added")
    if added_headers:
        ws_added.append(added_headers)
    for row_data in added_data:
        ws_added.append([str(v) if v is not None else "" for v in row_data])

    ws_removed = wb.create_sheet("Removed")
    if removed_headers:
        ws_removed.append(removed_headers)
    for row_data in removed_data:
        ws_removed.append([str(v) if v is not None else "" for v in row_data])

    ws_changed = wb.create_sheet("Changed")
    if changed_headers and changed_data:
        ws_changed.append([*key_columns, "Field", "Target", "Source"])
        for row_data in _iter_flatten_changed_rows(
            key_columns=key_columns,
            compare_columns=compare_columns,
            changed_headers=changed_headers,
            changed_data=changed_data,
            field_pairs=changed_field_pairs,
        ):
            ws_changed.append([str(v) if v is not None else "" for v in row_data])
    elif changed_headers:
        ws_changed.append(changed_headers)

    wb.save(file_path)

    report_id = f"rpt_{uuid.uuid4().hex[:8]}"
    summary_meta = {
        "added": added_count,
        "removed": removed_count,
        "changed": changed_count,
        "source_only_columns": schema_drift.get("source_only", []),
        "target_only_columns": schema_drift.get("target_only", []),
    }

    own = conn is None
    if own:
        conn = db.get_connection()
    db.create_report(
        conn,
        report_id=report_id,
        job_id=job_id,
        pair_id=pair_id,
        source_dataset=source_id,
        target_dataset=target_id,
        file_path=file_path,
        file_name=filename,
        summary=summary_meta,
    )
    if own:
        conn.close()

    return {
        "report_id": report_id,
        "file_path": file_path,
        "file_name": filename,
        "added": added_count,
        "removed": removed_count,
        "changed": changed_count,
    }


def write_comparison_report(
    comparison_result: Dict[str, Any],
    filename: Optional[str] = None,
    job_id: Optional[str] = None,
    pair_id: Optional[str] = None,
    conn=None,
) -> Dict[str, Any]:
    """Write a multi-tab XLSX report and register it in SQLite.

    ``comparison_result`` is the output of ``comparison.compare_full()``.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    if "error" in comparison_result:
        return {"error": comparison_result["error"]}

    source_id = comparison_result["source"]
    target_id = comparison_result["target"]
    key_columns = comparison_result.get("key_columns", [])
    compare_columns = comparison_result.get("compare_columns", [])
    schema_drift = comparison_result.get("schema_drift", {})
    added = comparison_result.get("added", {})
    removed = comparison_result.get("removed", {})
    changed = comparison_result.get("changed", {})

    # Counts
    added_count = len(added.get("data", []))
    removed_count = len(removed.get("data", []))
    changed_count = len(changed.get("data", []))
    if (added_count + removed_count + changed_count) >= _xlsx_stream_threshold():
        return _write_comparison_report_streaming(
            comparison_result=comparison_result,
            filename=filename,
            job_id=job_id,
            pair_id=pair_id,
            conn=conn,
        )

    # File path
    filename = _normalize_report_filename(
        filename=filename,
        default_stem=f"comparison_{source_id}_vs_{target_id}",
    )
    file_path = os.path.join(_reports_dir(conn), filename)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    removed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    changed_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    wb = Workbook()

    # ── Summary tab ──
    ws_sum = wb.active
    ws_sum.title = "Summary"
    summary_rows = [
        ("Comparison Report", ""),
        ("", ""),
        ("Source Dataset", source_id),
        ("Target Dataset", target_id),
        ("Key Columns", ", ".join(key_columns)),
        ("Compare Columns", ", ".join(compare_columns)),
        ("", ""),
        ("Timestamp", datetime.now(timezone.utc).isoformat()),
        ("", ""),
        ("Category", "Count"),
        ("Added (target-only)", added_count),
        ("Removed (source-only)", removed_count),
        ("Changed", changed_count),
        ("", ""),
        ("Schema Drift", ""),
        ("Source-only columns", ", ".join(schema_drift.get("source_only", []))),
        ("Target-only columns", ", ".join(schema_drift.get("target_only", []))),
    ]
    for row_data in summary_rows:
        ws_sum.append(list(row_data))
    # Style header
    ws_sum["A1"].font = Font(bold=True, size=14)
    for cell in ws_sum[10]:
        cell.font = header_font
        cell.fill = header_fill
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 60

    # ── Schema Drift tab ──
    drift_sheet_names = _unique_sheet_names(["Schema_Drift"])
    ws_drift = wb.create_sheet(drift_sheet_names[0])
    ws_drift.append(["Direction", "Column"])
    for cell in ws_drift[1]:
        cell.font = header_font
        cell.fill = header_fill
    for c in schema_drift.get("source_only", []):
        ws_drift.append(["Source only", c])
    for c in schema_drift.get("target_only", []):
        ws_drift.append(["Target only", c])
    ws_drift.column_dimensions["A"].width = 15
    ws_drift.column_dimensions["B"].width = 40

    # ── Added tab ──
    ws_added = wb.create_sheet("Added")
    added_headers = added.get("headers", [])
    added_data = added.get("data", [])
    if added_headers:
        ws_added.append(added_headers)
        for cell in ws_added[1]:
            cell.font = header_font
            cell.fill = header_fill
    for row_data in added_data:
        ws_added.append([str(v) if v is not None else "" for v in row_data])
        for cell in ws_added[ws_added.max_row]:
            cell.fill = added_fill

    # ── Removed tab ──
    ws_removed = wb.create_sheet("Removed")
    removed_headers = removed.get("headers", [])
    removed_data = removed.get("data", [])
    if removed_headers:
        ws_removed.append(removed_headers)
        for cell in ws_removed[1]:
            cell.font = header_font
            cell.fill = header_fill
    for row_data in removed_data:
        ws_removed.append([str(v) if v is not None else "" for v in row_data])
        for cell in ws_removed[ws_removed.max_row]:
            cell.fill = removed_fill

    # ── Changed tab ──
    ws_changed = wb.create_sheet("Changed")
    changed_headers = changed.get("headers", [])
    changed_data = changed.get("data", [])
    changed_field_pairs = changed.get("field_pairs", [])
    if changed_headers and changed_data:
        changed_headers, changed_data = _flatten_changed_rows(
            key_columns=key_columns,
            compare_columns=compare_columns,
            changed_headers=changed_headers,
            changed_data=changed_data,
            field_pairs=changed_field_pairs,
        )
    if changed_headers:
        ws_changed.append(changed_headers)
        for cell in ws_changed[1]:
            cell.font = header_font
            cell.fill = header_fill
    for row_data in changed_data:
        ws_changed.append([str(v) if v is not None else "" for v in row_data])
        for cell in ws_changed[ws_changed.max_row]:
            cell.fill = changed_fill

    # Auto-fit column widths (approximate)
    for ws in [ws_added, ws_removed, ws_changed]:
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells[:50]:  # sample first 50 rows
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(file_path)

    # Register in SQLite
    report_id = f"rpt_{uuid.uuid4().hex[:8]}"
    summary_meta = {
        "added": added_count,
        "removed": removed_count,
        "changed": changed_count,
        "source_only_columns": schema_drift.get("source_only", []),
        "target_only_columns": schema_drift.get("target_only", []),
    }

    own = conn is None
    if own:
        conn = db.get_connection()
    report = db.create_report(
        conn,
        report_id=report_id,
        job_id=job_id,
        pair_id=pair_id,
        source_dataset=source_id,
        target_dataset=target_id,
        file_path=file_path,
        file_name=filename,
        summary=summary_meta,
    )
    if own:
        conn.close()

    return {
        "report_id": report_id,
        "file_path": file_path,
        "file_name": filename,
        "added": added_count,
        "removed": removed_count,
        "changed": changed_count,
    }


# ═══════════════════════════════════════════════════════════════
#  Export query results to XLSX
# ═══════════════════════════════════════════════════════════════

def export_query_to_xlsx(
    headers: List[str],
    rows: Any,
    filename: Optional[str] = None,
    sql_query: Optional[str] = None,
    conn=None,
    job_id: Optional[str] = None,
) -> Dict[str, Any]:
    """Write arbitrary query results to an XLSX file."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    filename = _normalize_report_filename(
        filename=filename,
        default_stem="query_export",
    )
    file_path = os.path.join(_reports_dir(conn), filename)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    row_estimate: Optional[int] = None
    if hasattr(rows, "__len__"):
        try:
            row_estimate = int(len(rows))
        except Exception:
            row_estimate = None

    use_streaming = row_estimate is None or row_estimate >= _xlsx_stream_threshold()
    rows_written = 0
    if use_streaming:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Results")
        ws.append(headers)
        for row in _iter_rows(rows):
            ws.append([str(v) if v is not None else "" for v in row])
            rows_written += 1

        sql_sheet = wb.create_sheet("SQL")
        sql_sheet.append(["SQL Query"])
        sql_sheet.append([(sql_query or "").strip()])
        wb.save(file_path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in _iter_rows(rows):
            ws.append([str(v) if v is not None else "" for v in row])
            rows_written += 1

        # Auto-fit
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells[:50]:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        sql_sheet = wb.create_sheet("SQL")
        sql_sheet.append(["SQL Query"])
        sql_sheet["A1"].font = header_font
        sql_sheet["A1"].fill = header_fill
        sql_sheet["A2"] = (sql_query or "").strip()
        sql_sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        sql_sheet.column_dimensions["A"].width = 120

        wb.save(file_path)

    report_id = f"rpt_{uuid.uuid4().hex[:8]}"
    summary_meta = {
        "type": "query_export",
        "row_count": rows_written,
        "column_count": len(headers),
    }
    own = conn is None
    if own:
        conn = db.get_connection()
    db.create_report(
        conn,
        report_id=report_id,
        job_id=job_id,
        pair_id=None,
        source_dataset="query_export",
        target_dataset="query_export",
        file_path=file_path,
        file_name=filename,
        summary=summary_meta,
    )
    if own:
        conn.close()

    return {
        "report_id": report_id,
        "file_path": file_path,
        "file_name": filename,
        "row_count": rows_written,
    }


def export_column_summary_to_xlsx(
    summary_result: Dict[str, Any],
    top_n: int,
    filename: Optional[str] = None,
    conn=None,
) -> Dict[str, Any]:
    """Write column Top-N + blanks summary rows to an XLSX report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    dataset_id = summary_result.get("dataset", "dataset")
    summaries = summary_result.get("summaries", [])

    filename = _normalize_report_filename(
        filename=filename,
        default_stem=f"column_summary_{dataset_id}",
    )
    file_path = os.path.join(_reports_dir(conn), filename)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Dataset"
    ws["B1"] = dataset_id
    ws["A2"] = "Generated (UTC)"
    ws["B2"] = datetime.now(timezone.utc).isoformat()

    header_row = ["Column"] + [f"Top {i}" for i in range(1, top_n + 1)] + ["Blanks"]
    header_row_idx = 4
    for col_idx, title in enumerate(header_row, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, item in enumerate(summaries, start=header_row_idx + 1):
        column_name = item.get("column", "")
        ws.cell(row=row_idx, column=1, value=column_name)

        if item.get("error"):
            ws.cell(row=row_idx, column=2, value=f"ERROR: {item['error']}")
            for offset in range(1, top_n):
                ws.cell(row=row_idx, column=offset + 2, value="-")
            ws.cell(row=row_idx, column=len(header_row), value="-")
            continue

        top_values = item.get("top_values", [])
        for offset in range(top_n):
            if offset < len(top_values):
                val = top_values[offset].get("value")
                cnt = top_values[offset].get("count", 0)
                label = "NULL" if val is None else str(val)
                ws.cell(row=row_idx, column=offset + 2, value=f"{label} ({cnt})")
            else:
                ws.cell(row=row_idx, column=offset + 2, value="-")

        ws.cell(
            row=row_idx,
            column=len(header_row),
            value=item.get("blank_or_null_count", 0),
        )

    # Auto-fit width (sample top rows only for speed)
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells[:80]:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    wb.save(file_path)

    report_id = f"rpt_{uuid.uuid4().hex[:8]}"
    summary_meta = {
        "type": "column_summary",
        "dataset": dataset_id,
        "column_count": len(summaries),
        "top_n": top_n,
    }
    own = conn is None
    if own:
        conn = db.get_connection()
    db.create_report(
        conn,
        report_id=report_id,
        job_id=None,
        pair_id=None,
        source_dataset=dataset_id,
        target_dataset=dataset_id,
        file_path=file_path,
        file_name=filename,
        summary=summary_meta,
    )
    if own:
        conn.close()

    return {
        "report_id": report_id,
        "file_path": file_path,
        "file_name": filename,
        "dataset": dataset_id,
        "column_count": len(summaries),
        "top_n": top_n,
    }
